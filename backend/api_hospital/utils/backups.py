import csv
import io
import json
import re
import zipfile
from datetime import datetime
from pathlib import Path

from bson import ObjectId, json_util
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from utils.database import get_db

BASE_DIR = Path(__file__).resolve().parent.parent
BACKUP_DIR = BASE_DIR / 'backups'
AUTOMATION_CONFIG = BACKUP_DIR / 'automation.json'
BACKUP_CONTROL = BACKUP_DIR / 'control.json'
MAX_BACKUPS = 4
VALID_TYPES = {'completa', 'incremental', 'diferencial'}
VALID_FORMATS = {'json', 'csv', 'xlsx', 'pdf'}
RESTORABLE_FORMATS = {'json', 'csv', 'xlsx'}


class RestoreError(Exception):
    def __init__(self, restored, failed):
        self.restored = restored
        self.failed = failed
        details = '; '.join(
            f'{item["collection"]}: {item["error"]}' for item in failed
        )
        super().__init__(f'No se pudieron restaurar todas las colecciones. {details}')


def _safe_backup_path(filename):
    safe_name = Path(filename or '').name
    if safe_name != filename:
        raise ValueError('Nombre de respaldo inválido')
    path = BACKUP_DIR / safe_name
    if path.suffix.lower() not in {'.json', '.zip', '.xlsx', '.pdf'}:
        raise ValueError('Formato de respaldo inválido')
    if not path.is_file():
        raise FileNotFoundError('Respaldo no encontrado')
    return path


def _read_json(path, fallback):
    try:
        return json.loads(path.read_text(encoding='utf-8')) if path.exists() else fallback
    except (OSError, ValueError, TypeError):
        return fallback


def _write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')


def _control():
    return _read_json(BACKUP_CONTROL, {'ultima_completa': None, 'ultima_copia': None})


def _query_since(backup_type):
    if backup_type == 'completa':
        return {}
    control = _control()
    key = 'ultima_completa' if backup_type == 'diferencial' else 'ultima_copia'
    value = control.get(key)
    if not value:
        return {}
    since = datetime.fromisoformat(value)
    return {'$or': [
        {'updated_at': {'$gt': since}},
        {'fecha_modificacion': {'$gt': since}},
        {'_id': {'$gt': ObjectId.from_datetime(since)}},
    ]}


def _metadata(backup_type, backup_format, automatic, collections, counts):
    return {
        'version': 2,
        'database': get_db().name,
        'created_at': datetime.now().isoformat(),
        'type': backup_type,
        'format': backup_format,
        'automatic': automatic,
        'restorable': backup_format in RESTORABLE_FORMATS,
        'collections': collections,
        'document_counts': counts,
    }


def _readable_document(document):
    result = {}
    for key, value in document.items():
        if isinstance(value, (dict, list)):
            result[str(key)] = json_util.dumps(value, ensure_ascii=False)
        elif isinstance(value, (datetime, ObjectId)):
            result[str(key)] = str(value)
        elif value is None:
            result[str(key)] = ''
        else:
            result[str(key)] = str(value)
    return result


def _table_rows(documents):
    readable = [_readable_document(document) for document in documents]
    fields = sorted({field for document in readable for field in document})
    fields.append('_bson_document')
    rows = []
    for original, visible in zip(documents, readable):
        row = {field: visible.get(field, '') for field in fields}
        row['_bson_document'] = json_util.dumps(original, ensure_ascii=False)
        rows.append(row)
    return fields, rows


def _excel_sheet_name(collection_name, used_names):
    base = re.sub(r'[\[\]:*?/\\]', '_', collection_name).strip("'") or 'coleccion'
    base = base[:31]
    candidate = base
    suffix = 1
    while candidate.lower() in used_names:
        suffix_text = f'_{suffix}'
        candidate = f'{base[:31 - len(suffix_text)]}{suffix_text}'
        suffix += 1
    used_names.add(candidate.lower())
    return candidate


def obtener_colecciones():
    return sorted(get_db().list_collection_names())


def realizar_backup(tipo='completa', formato='json', colecciones=None, es_automatico=False):
    if tipo not in VALID_TYPES:
        raise ValueError('Tipo de respaldo inválido')
    if formato == 'excel':
        formato = 'xlsx'
    if formato not in VALID_FORMATS:
        raise ValueError('Formato de respaldo inválido')

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    db = get_db()
    available = set(db.list_collection_names())
    requested = set(colecciones) if colecciones else available
    selected = sorted(name for name in requested if isinstance(name, str) and name and not name.startswith('system.'))
    if not selected:
        raise ValueError('Selecciona al menos una colección válida')

    query = _query_since(tipo)
    collection_data = {name: list(db[name].find(query)) for name in selected}
    if tipo != 'completa':
        collection_data = {name: docs for name, docs in collection_data.items() if docs}
        if not collection_data:
            return None

    selected = list(collection_data)
    counts = {name: len(documents) for name, documents in collection_data.items()}
    metadata = _metadata(tipo, formato, es_automatico, selected, counts)
    label = 'automatico' if es_automatico else 'manual'
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    extension = 'zip' if formato == 'csv' else formato
    backup_path = BACKUP_DIR / f'backup_{label}_{tipo}_{timestamp}.{extension}'

    if formato == 'json':
        payload = {
            'metadata': metadata,
            'data': {name: documents for name, documents in collection_data.items()},
        }
        backup_path.write_text(json_util.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')

    elif formato == 'csv':
        with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as archive:
            archive.writestr('metadata.json', json.dumps(metadata, ensure_ascii=False, indent=2))
            for name, documents in collection_data.items():
                fields, rows = _table_rows(documents)
                buffer = io.StringIO()
                writer = csv.DictWriter(buffer, fieldnames=fields)
                writer.writeheader()
                writer.writerows(rows)
                archive.writestr(f'collections/{name}.csv', buffer.getvalue().encode('utf-8-sig'))

    elif formato == 'xlsx':
        workbook = Workbook()
        workbook.remove(workbook.active)
        meta_sheet = workbook.create_sheet('_INEO_METADATA')
        meta_sheet.sheet_state = 'hidden'
        sheet_map = {}
        used_names = {'_ineo_metadata'}
        for name, documents in collection_data.items():
            sheet_name = _excel_sheet_name(name, used_names)
            sheet_map[sheet_name] = name
            sheet = workbook.create_sheet(sheet_name)
            fields, rows = _table_rows(documents)
            sheet.append(fields)
            for row in rows:
                sheet.append([row.get(field, '') for field in fields])
        metadata['sheet_map'] = sheet_map
        meta_sheet['A1'] = json.dumps(metadata, ensure_ascii=False)
        workbook.save(backup_path)

    else:
        document = SimpleDocTemplate(str(backup_path), pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        story = [
            Paragraph('Copia de seguridad INEO', styles['Title']),
            Paragraph(f'Tipo: {tipo.upper()} · Fecha: {metadata["created_at"]}', styles['Normal']),
            Spacer(1, .2 * inch),
        ]
        for name, documents in collection_data.items():
            story.append(Paragraph(f'Colección: {name} ({len(documents)} documentos)', styles['Heading2']))
            visible = [_readable_document(item) for item in documents[:50]]
            fields = sorted({field for item in visible for field in item})[:8]
            if visible and fields:
                data = [fields] + [[str(item.get(field, ''))[:35] for field in fields] for item in visible]
                table = Table(data, repeatRows=1)
                table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), .4, colors.grey),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('FONTSIZE', (0, 0), (-1, -1), 6),
                ]))
                story.append(table)
            else:
                story.append(Paragraph('Sin datos', styles['Normal']))
            story.append(Spacer(1, .15 * inch))
        document.build(story)

    now = datetime.now().isoformat()
    control = _control()
    control['ultima_copia'] = now
    if tipo == 'completa':
        control['ultima_completa'] = now
    _write_json(BACKUP_CONTROL, control)
    limpiar_backups(MAX_BACKUPS)
    return str(backup_path)


def _load_backup(path):
    extension = path.suffix.lower()
    if extension == '.pdf':
        raise ValueError('Los respaldos PDF son solo de consulta y no se pueden restaurar')

    if extension == '.json':
        payload = json_util.loads(path.read_text(encoding='utf-8'))
        return payload.get('metadata', {}), payload.get('data', {})

    if extension == '.zip':
        data = {}
        with zipfile.ZipFile(path, 'r') as archive:
            names = archive.namelist()
            if any(Path(name).is_absolute() or '..' in Path(name).parts for name in names):
                raise ValueError('El respaldo contiene rutas no permitidas')
            metadata = json.loads(archive.read('metadata.json').decode('utf-8'))
            for name in names:
                if not name.startswith('collections/') or not name.endswith('.csv'):
                    continue
                collection = Path(name).stem
                content = archive.read(name).decode('utf-8-sig')
                rows = csv.DictReader(io.StringIO(content))
                data[collection] = [json_util.loads(row['_bson_document']) for row in rows if row.get('_bson_document')]
        return metadata, data

    workbook = load_workbook(path, read_only=True, data_only=True)
    metadata = json.loads(workbook['_INEO_METADATA']['A1'].value)
    data = {}
    data_sheets = [name for name in workbook.sheetnames if name != '_INEO_METADATA']
    sheet_map = metadata.get('sheet_map') or {}
    original_names = metadata.get('collections') or []
    for index, sheet_name in enumerate(data_sheets):
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows, []))
        if '_bson_document' not in headers:
            continue
        bson_index = headers.index('_bson_document')
        collection_name = sheet_map.get(sheet_name)
        if not collection_name and index < len(original_names):
            collection_name = original_names[index]
        collection_name = collection_name or sheet_name
        data[collection_name] = [
            json_util.loads(row[bson_index])
            for row in rows
            if len(row) > bson_index and row[bson_index]
        ]
    workbook.close()
    return metadata, data


def _restore_collection(db, collection_name, documents, complete):
    existed = collection_name in db.list_collection_names()
    collection = db[collection_name]
    previous_documents = list(collection.find({})) if existed else []

    try:
        if not existed:
            db.create_collection(collection_name)
            collection = db[collection_name]

        if complete:
            collection.delete_many({})
            if documents:
                collection.insert_many(
                    documents,
                    ordered=True,
                    bypass_document_validation=True,
                )
        else:
            for document in documents:
                if '_id' in document:
                    collection.replace_one(
                        {'_id': document['_id']},
                        document,
                        upsert=True,
                        bypass_document_validation=True,
                    )
                else:
                    collection.insert_one(
                        document,
                        bypass_document_validation=True,
                    )
    except Exception:
        collection.delete_many({})
        if previous_documents:
            collection.insert_many(
                previous_documents,
                ordered=True,
                bypass_document_validation=True,
            )
        elif not existed:
            collection.drop()
        raise


def restaurar_backup(filename):
    backup_path = _safe_backup_path(filename)
    metadata, pending = _load_backup(backup_path)
    if not pending:
        raise ValueError('El respaldo no contiene colecciones')

    db = get_db()
    realizar_backup(
        tipo='completa',
        formato='json',
        colecciones=list(pending),
        es_automatico=True,
    )

    complete = metadata.get('type', 'completa') == 'completa'
    restored = []
    failed = []
    for collection_name, documents in pending.items():
        try:
            _restore_collection(db, collection_name, documents, complete)
            restored.append(collection_name)
        except Exception as exc:
            failed.append({
                'collection': collection_name,
                'error': str(exc),
            })

    if failed:
        raise RestoreError(restored, failed)
    return restored


def _file_metadata(path):
    try:
        metadata, _ = _load_backup(path)
        return metadata
    except (ValueError, KeyError, OSError, json.JSONDecodeError, zipfile.BadZipFile):
        parts = path.stem.split('_')
        backup_type = next((item for item in parts if item in VALID_TYPES), 'completa')
        return {
            'type': backup_type,
            'format': 'pdf' if path.suffix.lower() == '.pdf' else path.suffix.lstrip('.'),
            'automatic': 'automatico' in parts,
            'restorable': path.suffix.lower() != '.pdf',
        }


def list_backups():
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    result = []
    for path in BACKUP_DIR.iterdir():
        if (
            not path.is_file()
            or not path.name.startswith('backup_')
            or path.suffix.lower() not in {'.json', '.zip', '.xlsx', '.pdf'}
        ):
            continue
        metadata = _file_metadata(path)
        stat = path.stat()
        result.append({
            'filename': path.name,
            'size': stat.st_size,
            'date': metadata.get('created_at') or datetime.fromtimestamp(stat.st_mtime).isoformat(),
            'automatic': bool(metadata.get('automatic', 'automatico' in path.name)),
            'type': metadata.get('type', 'completa'),
            'format': metadata.get('format', path.suffix.lstrip('.')),
            'restorable': bool(metadata.get('restorable', path.suffix.lower() != '.pdf')),
            'collections': metadata.get('collections', []),
        })
    return sorted(result, key=lambda item: item['date'], reverse=True)


def eliminar_backup(filename):
    _safe_backup_path(filename).unlink()


def limpiar_backups(keep=MAX_BACKUPS):
    keep = max(1, min(int(keep), 50))
    groups = {}
    for backup in list_backups():
        group = (backup['automatic'], backup['type'])
        groups.setdefault(group, []).append(backup)
    for backups in groups.values():
        for backup in backups[keep:]:
            (BACKUP_DIR / backup['filename']).unlink(missing_ok=True)


def cargar_config_automatizacion():
    defaults = {
        'activo': False,
        'tipo': 'completa',
        'formato': 'json',
        'intervalo': 1440,
        'colecciones': [],
        'max_backups': MAX_BACKUPS,
    }
    return {**defaults, **_read_json(AUTOMATION_CONFIG, {})}


def guardar_config_automatizacion(config):
    _write_json(AUTOMATION_CONFIG, config)
    return config


def check_db_health():
    try:
        get_db().command('ping')
        return {'status': 'ok', 'message': 'MongoDB conectado'}
    except Exception as exc:
        return {'status': 'error', 'message': str(exc)}
